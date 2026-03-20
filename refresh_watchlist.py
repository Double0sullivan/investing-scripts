#!/usr/bin/env python3
"""
Watchlist Tracker — Live Price Refresher
-----------------------------------------
Reads tickers from your Watchlist_Tracker.xlsx spreadsheet,
fetches live data from Yahoo Finance, and writes it all back.

Usage:
    py refresh_watchlist.py

Requirements:
    pip install yfinance openpyxl
"""

import sys
import datetime
import io
import subprocess

def ensure(pkg, import_as=None):
    try:
        __import__(import_as or pkg)
    except ImportError:
        print(f"Installing {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

ensure("yfinance")
ensure("matplotlib")

import yfinance as yf
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image as XLImage

# ── Config ─────────────────────────────────────────────────────────────────────
SPREADSHEET  = "Watchlist_Tracker.xlsx"
SHEET_NAME   = "Watchlist"
TICKER_COL   = 2    # Column B
DATA_START_ROW = 7
LAST_UPDATED_CELL = "B4"

# Column positions for data output
COL_NAME      = 3   # C - Company Name
COL_PRICE     = 4   # D - Price
COL_CHANGE    = 5   # E - Change Today ($)
COL_CHANGEPCT = 6   # F - % Change Today
COL_52H       = 7   # G - 52W High
COL_52H_PCT   = 8   # H - % from 52W High
COL_52L       = 9   # I - 52W Low
COL_52L_PCT   = 10  # J - % from 52W Low
COL_MCAP      = 11  # K - Market Cap
COL_PE        = 12  # L - P/E Ratio
COL_DIV       = 13  # M - Dividend Yield

C_GREEN = "C8E6C9"
C_RED   = "FFCDD2"
C_WHITE = "FFFFFF"
C_GREY  = "F0F4F8"

def fmt_mcap(val):
    if val is None: return "—"
    if val >= 1_000_000_000_000:
        return f"{val/1_000_000_000_000:.2f}T"
    if val >= 1_000_000_000:
        return f"{val/1_000_000_000:.1f}B"
    if val >= 1_000_000:
        return f"{val/1_000_000:.1f}M"
    return str(val)

def get_data(ticker_symbol):
    t = yf.Ticker(ticker_symbol)
    info = t.info

    # Price - try fast_info first, fallback to info
    try:
        price = t.fast_info.last_price
    except:
        price = info.get("currentPrice") or info.get("regularMarketPrice")

    prev_close  = info.get("previousClose") or info.get("regularMarketPreviousClose")
    change      = (price - prev_close) if price and prev_close else None
    change_pct  = (change / prev_close) if change and prev_close else None

    high_52w    = info.get("fiftyTwoWeekHigh")
    low_52w     = info.get("fiftyTwoWeekLow")
    pct_52h     = ((price - high_52w) / high_52w) if price and high_52w else None
    pct_52l     = ((price - low_52w)  / low_52w)  if price and low_52w  else None

    mcap        = info.get("marketCap")
    pe          = info.get("trailingPE") or info.get("forwardPE")
    div_yield   = info.get("dividendYield")
    name        = info.get("shortName") or info.get("longName") or ticker_symbol

    return {
        "name":       name,
        "price":      price,
        "change":     change,
        "change_pct": change_pct,
        "high_52w":   high_52w,
        "pct_52h":    pct_52h,
        "low_52w":    low_52w,
        "pct_52l":    pct_52l,
        "mcap":       mcap,
        "pe":         pe,
        "div_yield":  div_yield,
    }

def write_cell(ws, row, col, value, num_format=None, bold=False, color=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if num_format:
        cell.number_format = num_format
    if bold:
        cell.font = Font(name="Arial", bold=True, size=10,
                         color=color or "0D1B2A")
    if color and not bold:
        cell.font = Font(name="Arial", size=10, color=color)

def main():
    print("=" * 55)
    print("  Watchlist Tracker — Live Refresh")
    print("=" * 55)

    try:
        wb = load_workbook(SPREADSHEET)
    except FileNotFoundError:
        print(f"\n  ERROR: '{SPREADSHEET}' not found.")
        print("  Make sure this script is in the same folder as the spreadsheet.")
        sys.exit(1)

    ws = wb[SHEET_NAME]

    # Read all tickers from column B
    tickers = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        val = ws.cell(row=row, column=TICKER_COL).value
        if val and str(val).strip():
            tickers.append((row, str(val).strip().upper()))

    if not tickers:
        print("\n  No tickers found in column B. Add some tickers and try again.")
        sys.exit(0)

    print(f"\n  Found {len(tickers)} ticker(s). Fetching data...\n")

    ok, failed = 0, 0

    for row, ticker in tickers:
        bg = C_GREY if (row - DATA_START_ROW) % 2 == 0 else C_WHITE
        try:
            d = get_data(ticker)

            ws.cell(row=row, column=COL_NAME).value      = d["name"]
            ws.cell(row=row, column=COL_PRICE).value     = d["price"]
            ws.cell(row=row, column=COL_PRICE).number_format = "#,##0.00"

            ws.cell(row=row, column=COL_CHANGE).value    = d["change"]
            ws.cell(row=row, column=COL_CHANGE).number_format = '+#,##0.00;(#,##0.00);"-"'

            ws.cell(row=row, column=COL_CHANGEPCT).value = d["change_pct"]
            ws.cell(row=row, column=COL_CHANGEPCT).number_format = '+0.00%;(0.00%);"-"'

            ws.cell(row=row, column=COL_52H).value       = d["high_52w"]
            ws.cell(row=row, column=COL_52H).number_format = "#,##0.00"

            ws.cell(row=row, column=COL_52H_PCT).value   = d["pct_52h"]
            ws.cell(row=row, column=COL_52H_PCT).number_format = '0.00%;(0.00%);"-"'

            ws.cell(row=row, column=COL_52L).value       = d["low_52w"]
            ws.cell(row=row, column=COL_52L).number_format = "#,##0.00"

            ws.cell(row=row, column=COL_52L_PCT).value   = d["pct_52l"]
            ws.cell(row=row, column=COL_52L_PCT).number_format = '+0.00%;(0.00%);"-"'

            ws.cell(row=row, column=COL_MCAP).value      = fmt_mcap(d["mcap"])

            pe_val = round(d["pe"], 1) if d["pe"] else None
            ws.cell(row=row, column=COL_PE).value        = pe_val if pe_val else "—"
            if pe_val:
                ws.cell(row=row, column=COL_PE).number_format = "0.0"

            div = d["div_yield"]
            # yfinance returns dividendYield as decimal e.g. 0.0042 = 0.42%
            # Write as pre-formatted string to avoid Excel doubling the percentage
            div_str = f"{div * 100:.2f}%" if div else "—"
            ws.cell(row=row, column=COL_DIV).value = div_str
            ws.cell(row=row, column=COL_DIV).number_format = "@"

            # Colour % change cell green/red
            chg_cell = ws.cell(row=row, column=COL_CHANGEPCT)
            if d["change_pct"] is not None:
                color = C_GREEN if d["change_pct"] >= 0 else C_RED
                chg_cell.fill = PatternFill("solid", fgColor=color)

            pct = f"{d['change_pct']*100:+.2f}%" if d["change_pct"] else "—"
            print(f"  ✓ {ticker:<8}  {d['name'][:28]:<28}  ${d['price']:>10,.2f}  {pct}")
            ok += 1

        except Exception as e:
            print(f"  ✗ {ticker:<8}  ERROR: {e}")
            for col in range(COL_NAME, COL_DIV + 1):
                ws.cell(row=row, column=col).value = "ERROR"
            failed += 1

    # Timestamp
    now = datetime.datetime.now().strftime("%B %d, %Y at %I:%M %p")
    ws[LAST_UPDATED_CELL] = f"Last updated: {now}"

    # ── Generate price chart ───────────────────────────────────────────────
    chart_generated = False
    if "Chart" in wb.sheetnames:
        wc = wb["Chart"]
        ticker_val = wc["C7"].value
        period_val = wc["C8"].value

        if ticker_val and str(ticker_val).strip():
            chart_ticker = str(ticker_val).strip().upper()
            try:
                months = int(float(str(period_val))) if period_val else 12
            except:
                months = 12

            print(f"  Generating chart for {chart_ticker} ({months} months)...")
            try:
                # Build yfinance period string
                if months <= 1:
                    period_str = "1mo"
                elif months <= 3:
                    period_str = "3mo"
                elif months <= 6:
                    period_str = "6mo"
                elif months <= 12:
                    period_str = "1y"
                elif months <= 24:
                    period_str = "2y"
                elif months <= 60:
                    period_str = "5y"
                else:
                    period_str = "10y"

                hist = yf.Ticker(chart_ticker).history(period=period_str)

                if not hist.empty:
                    fig, ax = plt.subplots(figsize=(12, 5))
                    fig.patch.set_facecolor("#0D1B2A")
                    ax.set_facecolor("#0D1B2A")

                    prices = hist["Close"]
                    dates  = hist.index

                    color = "#5FA8D3"
                    ax.plot(dates, prices, color=color, linewidth=1.8, zorder=3)
                    ax.fill_between(dates, prices, prices.min(), alpha=0.15, color=color)

                    # Style axes
                    ax.spines["top"].set_visible(False)
                    ax.spines["right"].set_visible(False)
                    ax.spines["left"].set_color("#334466")
                    ax.spines["bottom"].set_color("#334466")
                    ax.tick_params(colors="#AABBCC", labelsize=9)
                    ax.yaxis.label.set_color("#AABBCC")
                    ax.xaxis.label.set_color("#AABBCC")
                    ax.grid(True, color="#1B3050", linewidth=0.5, zorder=0)

                    # Title
                    start_p = float(prices.iloc[0])
                    end_p   = float(prices.iloc[-1])
                    chg_pct = (end_p - start_p) / start_p * 100
                    chg_color = "#2DC653" if chg_pct >= 0 else "#E63946"
                    sign = "+" if chg_pct >= 0 else ""
                    fig.suptitle(
                        f"{chart_ticker}  —  {end_p:,.2f}  ({sign}{chg_pct:.1f}%  over {months} months)",
                        color=chg_color, fontsize=13, fontweight="bold", y=0.97
                    )

                    # Date formatting
                    if months <= 3:
                        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d %b"))
                    elif months <= 24:
                        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
                    else:
                        ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
                    plt.xticks(rotation=30, ha="right")

                    # Price axis formatting
                    ax.yaxis.set_major_formatter(
                        matplotlib.ticker.FuncFormatter(lambda x, _: f"{x:,.0f}")
                    )

                    plt.tight_layout(rect=[0, 0, 1, 0.94])

                    # Save chart to bytes
                    buf = io.BytesIO()
                    plt.savefig(buf, format="png", dpi=130,
                                facecolor=fig.get_facecolor(), bbox_inches="tight")
                    plt.close(fig)
                    buf.seek(0)

                    # Remove old images from chart sheet
                    wc._images = []

                    # Embed into spreadsheet at row 12
                    img = XLImage(buf)
                    img.anchor = "B12"
                    wc.add_image(img)

                    # Update timestamp on chart sheet
                    wc["B4"] = f"Last chart generated: {chart_ticker}  |  {months} months  |  {now}"

                    chart_generated = True
                    print(f"  ✓ Chart embedded for {chart_ticker}")
                else:
                    print(f"  ✗ No price history found for {chart_ticker}")

            except Exception as e:
                print(f"  ✗ Chart error: {e}")

    wb.save(SPREADSHEET)

    print()
    print(f"  ✓ Done! {ok} updated, {failed} failed.")
    if chart_generated:
        print(f"  ✓ Chart updated in the Chart tab.")
    print(f"  ✓ Saved: {SPREADSHEET}")
    print(f"  ✓ Time:  {now}")
    print("=" * 55)

if __name__ == "__main__":
    main()
