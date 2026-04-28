#!/usr/bin/env python3
"""
Watchlist — Per-Stock Chart Generator
--------------------------------------
- Reads tickers from the Watchlist tab
- Reads indices from the Indices tab
- Fetches 5-year history for stocks and indices from Yahoo Finance
- Stores index data locally so re-runs are fast
- Creates one chart tab per stock
- Type an index into ANY stock tab's yellow cell C6 — it applies to ALL tabs

Usage:  py build_stock_charts.py
Needs:  py -m pip install yfinance matplotlib openpyxl pandas
"""

import sys, io, datetime, subprocess, warnings
warnings.filterwarnings("ignore")

def ensure(pkg):
    try: __import__(pkg)
    except ImportError:
        print(f"  Installing {pkg}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

for p in ["yfinance", "matplotlib", "pandas", "openpyxl"]:
    ensure(p)

import yfinance as yf
import pandas as pd
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.dates as mdates
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ── Config ────────────────────────────────────────────────────────────────────
SPREADSHEET      = "Watchlist_Tracker.xlsx"
WATCHLIST_SHEET  = "Watchlist"
INDICES_SHEET    = "Indices"
TICKER_COL       = 2
DATA_START_ROW   = 7
PERIOD           = "5y"
DEFAULT_INDEX    = "^GSPC"   # ← Change this to your preferred index
INDEX_DATA_ROW   = 32   # row where index price history is stored in Indices tab

C_DARK   = "0D1B2A"
C_MID    = "1B4965"
C_LIGHT  = "CAE9FF"
C_WHITE  = "FFFFFF"
C_GREY   = "F0F4F8"
C_YELLOW = "FFFDE7"
BG       = "#0D1B2A"

DEFAULT_INDICES = ["^GSPC", "^IXIC", "^FTSE", "^DJI", "^RUT"]

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(c):    return PatternFill("solid", fgColor=c)
def xfont(bold=False, color=C_DARK, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def xalign(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=False)
def xborder():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def set_cell(ws, ref, value, bold=False, color=C_DARK, size=10,
             italic=False, bg=None, h="left", bordered=False):
    c = ws[ref]
    c.value     = value
    c.font      = xfont(bold=bold, color=color, size=size, italic=italic)
    c.alignment = xalign(h)
    if bg:      c.fill   = fill(bg)
    if bordered: c.border = xborder()
    return c

# ── Indices tab ───────────────────────────────────────────────────────────────
def ensure_indices_sheet(wb):
    if INDICES_SHEET in wb.sheetnames:
        return
    ws = wb.create_sheet(INDICES_SHEET, 1)
    ws.sheet_view.showGridLines = False
    for col, w in [("A",2),("B",18),("C",28)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[6].height = 24
    ws.merge_cells("B2:C2")
    set_cell(ws, "B2", "Indices", bold=True, color=C_WHITE, size=16, bg=C_DARK, h="center")
    ws.merge_cells("B3:C3")
    set_cell(ws, "B3", "Add index tickers in column B below. Run build_stock_charts.py to fetch data.",
             italic=True, color=C_WHITE, size=10, bg=C_MID, h="center")
    ws.merge_cells("B4:C4")
    set_cell(ws, "B4", "Last updated: —", italic=True, color="888888", size=9, bg=C_GREY, h="right")
    for col, label in [("B","Index Ticker"), ("C","Full Name")]:
        set_cell(ws, f"{col}6", label, bold=True, color=C_WHITE, size=10, bg=C_MID, h="center", bordered=True)
    for i, idx in enumerate(DEFAULT_INDICES):
        r = 7 + i
        ws.row_dimensions[r].height = 20
        set_cell(ws, f"B{r}", idx, bold=True, color="0D47A1", size=10, bg=C_YELLOW, h="center", bordered=True)
        ws[f"C{r}"].fill   = fill(C_GREY if i%2==0 else C_WHITE)
        ws[f"C{r}"].border = xborder()
    for i in range(len(DEFAULT_INDICES), 20):
        r = 7 + i
        ws.row_dimensions[r].height = 20
        ws[f"B{r}"].fill   = fill(C_YELLOW)
        ws[f"B{r}"].border = xborder()
        ws[f"B{r}"].alignment = xalign("center")
        ws[f"C{r}"].fill   = fill(C_GREY if i%2==0 else C_WHITE)
        ws[f"C{r}"].border = xborder()

def read_index_tickers(wb):
    ws = wb[INDICES_SHEET]
    return [
        str(ws[f"B{r}"].value).strip().upper()
        for r in range(7, 27)
        if ws[f"B{r}"].value and str(ws[f"B{r}"].value).strip()
        and str(ws[f"B{r}"].value).strip() != "Index Ticker"
    ]

def fetch_and_store_index_data(wb, indices):
    ws = wb[INDICES_SHEET]
    if not indices:
        return pd.DataFrame()
    print(f"\n  Fetching index history: {', '.join(indices)}")
    all_data = {}
    for idx in indices:
        try:
            t    = yf.Ticker(idx)
            hist = t.history(period=PERIOD)["Close"]
            if hist.empty:
                print(f"  ! {idx} — no data"); continue
            hist.index = pd.to_datetime(hist.index).tz_localize(None).normalize()
            all_data[idx] = hist
            name = (t.info.get("shortName") or t.info.get("longName") or idx)
            for r in range(7, 27):
                if ws[f"B{r}"].value and str(ws[f"B{r}"].value).strip().upper() == idx:
                    ws[f"C{r}"] = name
                    ws[f"C{r}"].font = xfont(color=C_DARK, size=10)
                    break
            print(f"  + {idx}  ({len(hist)} days)")
        except Exception as e:
            print(f"  ! {idx} — {e}")
    if not all_data:
        return pd.DataFrame()
    df = pd.DataFrame(all_data).sort_index()
    # Store header
    hdr = INDEX_DATA_ROW
    ws.cell(row=hdr,   column=2).value = "Date"
    ws.cell(row=hdr,   column=2).font  = xfont(bold=True, color=C_WHITE, size=8)
    ws.cell(row=hdr,   column=2).fill  = fill(C_MID)
    from openpyxl.utils import get_column_letter
    col_map = {}
    for i, idx in enumerate(df.columns):
        col = 3 + i
        col_map[idx] = col
        ws.cell(row=hdr, column=col).value = idx
        ws.cell(row=hdr, column=col).font  = xfont(bold=True, color=C_WHITE, size=8)
        ws.cell(row=hdr, column=col).fill  = fill(C_MID)
        ws.column_dimensions[get_column_letter(col)].width = 13
    for r_off, (date, row_data) in enumerate(df.iterrows()):
        r = hdr + 1 + r_off
        ws.cell(row=r, column=2).value = date.strftime("%Y-%m-%d")
        ws.cell(row=r, column=2).font  = xfont(color="AAAAAA", size=7)
        for idx, col in col_map.items():
            val = row_data.get(idx)
            if pd.notna(val):
                ws.cell(row=r, column=col).value          = round(float(val), 4)
                ws.cell(row=r, column=col).number_format  = "#,##0.00"
                ws.cell(row=r, column=col).font           = xfont(color=C_DARK, size=7)
    ws["B4"] = f"Last updated: {datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    return df

def load_index_data_from_sheet(wb):
    ws  = wb[INDICES_SHEET]
    hdr = INDEX_DATA_ROW
    col_map = {}
    for col in range(3, 30):
        v = ws.cell(row=hdr, column=col).value
        if v and str(v).strip():
            col_map[col] = str(v).strip().upper()
        elif col > 5 and not v:
            break
    if not col_map:
        return pd.DataFrame()
    rows = []
    for r in range(hdr + 1, hdr + 2000):
        dv = ws.cell(row=r, column=2).value
        if not dv: break
        try:
            row = {"Date": pd.to_datetime(str(dv))}
        except Exception:
            break
        for col, ticker in col_map.items():
            row[ticker] = ws.cell(row=r, column=col).value
        rows.append(row)
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows).set_index("Date")
    df.index = pd.to_datetime(df.index).normalize()
    return df

# ── Chart builder ─────────────────────────────────────────────────────────────
def style_ax(ax):
    ax.set_facecolor(BG)
    for sp in ["top","right"]:   ax.spines[sp].set_visible(False)
    for sp in ["left","bottom"]: ax.spines[sp].set_color("#1B4965")
    ax.tick_params(colors="#8AAABB", labelsize=9)
    ax.grid(True, color="#112233", linewidth=0.6, zorder=0)
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b '%y"))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha="right")

def build_chart(ticker, stock_series, index_ticker, index_df):
    idx_key  = index_ticker.strip().upper() if index_ticker else None
    show_rel = bool(idx_key and not index_df.empty and idx_key in index_df.columns)

    s = stock_series.copy()
    s.index = s.index.normalize()

    if show_rel:
        ix     = index_df[idx_key].dropna()
        common = s.index.intersection(ix.index)
        s_c    = s.loc[common]
        i_c    = ix.loc[common]
        rel    = (s_c / s_c.iloc[0]) / (i_c / i_c.iloc[0])
    else:
        s_c = s

    nrows = 2 if show_rel else 1
    fig, axes = plt.subplots(nrows, 1, figsize=(13, 5.5*nrows),
                             facecolor=BG, gridspec_kw={"hspace":0.5})
    if nrows == 1:
        axes = [axes]

    # Panel 1 — absolute price
    ax1 = axes[0]
    style_ax(ax1)
    sp, ep   = float(s_c.iloc[0]), float(s_c.iloc[-1])
    chg      = (ep - sp) / sp * 100
    lc       = "#2DC653" if chg >= 0 else "#E63946"
    sign     = "+" if chg >= 0 else ""
    ax1.plot(s_c.index, s_c.values, color=lc, linewidth=1.8, zorder=3)
    ax1.fill_between(s_c.index, s_c.values, s_c.min(), alpha=0.12, color=lc)
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x:,.2f}"))
    ax1.set_title(f"{ticker}  —  {ep:,.2f}  ({sign}{chg:.1f}%  over 5 years)",
                  color=lc, fontsize=13, fontweight="bold", pad=10, loc="left")
    ax1.set_ylabel("Price", color="#8AAABB", fontsize=9)
    ax1.annotate(f"{ep:,.2f}", xy=(s_c.index[-1], ep),
                 xytext=(8,0), textcoords="offset points",
                 color=lc, fontsize=9, fontweight="bold", va="center")

    # Panel 2 — relative
    if show_rel:
        ax2 = axes[1]
        style_ax(ax2)
        re    = float(rel.iloc[-1])
        rchg  = (re - 1) * 100
        rc    = "#2DC653" if rchg >= 0 else "#E63946"
        rs    = "+" if rchg >= 0 else ""
        ax2.plot(rel.index, rel.values, color=rc, linewidth=1.8, zorder=3)
        ax2.fill_between(rel.index, rel.values, rel.min(), alpha=0.12, color=rc)
        ax2.axhline(1.0, color="#334455", linewidth=0.9, linestyle="--", zorder=2)
        ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x:.2f}x"))
        ax2.set_title(f"Relative to {idx_key}  —  {rs}{rchg:.1f}%  vs index over 5 years",
                      color=rc, fontsize=11, fontweight="bold", pad=10, loc="left")
        ax2.set_ylabel("Ratio  (1.0 = same as index)", color="#8AAABB", fontsize=9)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=130, facecolor=BG, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf

# ── Stock sheet builder ───────────────────────────────────────────────────────
def make_stock_sheet(wb, ticker, index_ticker, available_indices):
    name = ticker[:31]
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for col, w in [("A",2),("B",22),("C",18),("D",44),("E",2)]:
        ws.column_dimensions[col].width = w
    for r, h in [(2,30),(3,20),(4,20),(6,22)]:
        ws.row_dimensions[r].height = h
    for r in range(7, 70):
        ws.row_dimensions[r].height = 18

    ws.merge_cells("B2:D2")
    set_cell(ws, "B2", f"  {ticker}  —  5-Year Price History",
             bold=True, color=C_WHITE, size=14, bg=C_DARK, h="center")
    ws.merge_cells("B3:D3")
    set_cell(ws, "B3",
             "Type an index in the yellow cell and re-run build_stock_charts.py for relative performance.",
             italic=True, color=C_WHITE, size=9, bg=C_MID, h="center")
    ws.merge_cells("B4:D4")
    set_cell(ws, "B4",
             f"Last updated: {datetime.datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
             italic=True, color="888888", size=9, bg=C_GREY, h="right")

    set_cell(ws, "B6", "Compare vs Index",
             bold=True, color=C_DARK, size=10, bg=C_GREY, h="left", bordered=True)

    # ── Yellow input cell C6 ──────────────────────────────────────────────────
    c6 = ws["C6"]
    c6.value     = index_ticker if index_ticker else ""
    c6.font      = xfont(bold=True, color="0D47A1", size=11)
    c6.fill      = fill(C_YELLOW)
    c6.alignment = xalign("center")
    c6.border    = xborder()

    hint = "  Available: " + "  |  ".join(available_indices[:8])
    set_cell(ws, "D6", hint, italic=True, color="888888", size=9, bg=C_WHITE, h="left", bordered=True)

    return ws

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  Watchlist — Per-Stock Chart Generator")
    print("=" * 60)

    try:
        wb = load_workbook(SPREADSHEET)
    except FileNotFoundError:
        print(f"\n  ERROR: '{SPREADSHEET}' not found.")
        sys.exit(1)

    ensure_indices_sheet(wb)

    # ── Step 1: read any index already saved in a stock tab BEFORE any reload ──
    ws_main = wb[WATCHLIST_SHEET]
    tickers = [
        str(ws_main.cell(row=r, column=TICKER_COL).value).strip().upper()
        for r in range(DATA_START_ROW, ws_main.max_row + 1)
        if ws_main.cell(row=r, column=TICKER_COL).value
        and str(ws_main.cell(row=r, column=TICKER_COL).value).strip()
    ]

    if not tickers:
        print("\n  No tickers found in Watchlist tab.")
        sys.exit(0)

    # Find default index from any existing stock tab
    default_index = None
    for ticker in tickers:
        sname = ticker[:31]
        if sname in wb.sheetnames:
            v = wb[sname]["C6"].value
            if v and str(v).strip():
                default_index = str(v).strip().upper()
                print(f"\n  Default index: {default_index}  (read from {ticker} tab)")
                break

    if not default_index:
        print("\n  No index set yet — charts will show absolute price only.")
        print("  Tip: type an index (e.g. ^GSPC) into the yellow C6 cell on any stock tab, then re-run.")

    # ── Step 2: fetch and store index data ────────────────────────────────────
    indices  = read_index_tickers(wb)
    print(f"\n  Indices tab: {', '.join(indices) if indices else 'none'}")
    fetch_and_store_index_data(wb, indices)

    # Save once so index data is on disk, then reload cleanly
    wb.save(SPREADSHEET)
    wb = load_workbook(SPREADSHEET)

    index_df  = load_index_data_from_sheet(wb)
    available = list(index_df.columns) if not index_df.empty else indices

    print(f"\n  Building charts for: {', '.join(tickers)}\n")
    ok, failed = 0, 0

    for ticker in tickers:
        sname = ticker[:31]

        # Each stock uses default_index unless its own tab has a different one
        index_ticker = default_index
        if sname in wb.sheetnames:
            v = wb[sname]["C6"].value
            if v and str(v).strip():
                index_ticker = str(v).strip().upper()

        print(f"  {ticker}  →  index: {index_ticker or 'none'}")

        try:
            hist = yf.Ticker(ticker).history(period=PERIOD)["Close"]
            if hist.empty:
                raise ValueError("No data returned")
            hist.index = pd.to_datetime(hist.index).tz_localize(None).normalize()

            buf = build_chart(ticker, hist, index_ticker, index_df)
            ws  = make_stock_sheet(wb, ticker, index_ticker, available)

            # Belt-and-braces: set C6 again explicitly after sheet creation
            ws["C6"].value     = index_ticker if index_ticker else ""
            ws["C6"].font      = xfont(bold=True, color="0D47A1", size=11)
            ws["C6"].fill      = fill(C_YELLOW)
            ws["C6"].alignment = xalign("center")
            ws["C6"].border    = xborder()

            img        = XLImage(buf)
            img.anchor = "B8"
            ws.add_image(img)
            print(f"  + {ticker}  OK")
            ok += 1

        except Exception as e:
            print(f"  ! {ticker}  — {e}")
            failed += 1

    wb.save(SPREADSHEET)
    print()
    print(f"  Done:  {ok} chart(s) built,  {failed} failed.")
    print(f"  Saved: {SPREADSHEET}")
    print()
    print("  To compare vs an index:")
    print("  1. Open the spreadsheet, go to any stock tab")
    print("  2. Type an index into the yellow C6 cell (e.g. ^GSPC, ^IXIC, ^FTSE)")
    print("  3. Close the file, re-run:  py build_stock_charts.py")
    print("  That index will be applied to ALL stock tabs automatically.")
    print("=" * 60)

if __name__ == "__main__":
    main()
